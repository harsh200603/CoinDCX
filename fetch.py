import requests
import pandas as pd
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ====== CONFIG ======
coins = ["BTCUSDT", "ETHUSDT", "XRPUSDT", "BNBUSDT", "DOGEUSDT"]
interval = 3  # seconds
output_file = "crypto_top5.xlsx"

def get_price(symbol):
    """Fetch latest price for given symbol from CoinDCX."""
    url = "https://api.coindcx.com/exchange/ticker"
    try:
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        data = response.json()
        for ticker in data:
            if ticker["market"] == symbol:
                return float(ticker["last_price"])
        return None
    except Exception as e:
        print(f"Error fetching data for {symbol}:", e)
        return None

if __name__ == "__main__":
    # Store all coin data
    data_dict = {coin: [] for coin in coins}

    print(f"Tracking {', '.join(coins)} every {interval} seconds... Press Ctrl+C to stop.")

    try:
        while True:
            for coin in coins:
                price = get_price(coin)
                if price is not None:
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    data_dict[coin].append((price, timestamp))
                    print(f"{coin} | Price: {price} | Time: {timestamp}")

            # Prepare data for Excel
            excel_data = {}
            for coin in coins:
                prices = [p for p, _ in data_dict[coin]]
                times = [t for _, t in data_dict[coin]]
                excel_data[f"{coin} Price"] = prices
                excel_data[f"{coin} Time"] = times

            df = pd.DataFrame(excel_data)
            df.to_excel(output_file, index=False)

            # Open and format with openpyxl
            wb = load_workbook(output_file)
            ws = wb.active

            # Merge and style headers
            col = 1
            for coin in coins:
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
                ws.cell(row=1, column=col).value = coin
                ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=2, column=col).value = "Price"
                ws.cell(row=2, column=col+1).value = "Time"
                col += 2

            wb.save(output_file)

            time.sleep(interval)

    except KeyboardInterrupt:
        print("\nStopped tracking.")